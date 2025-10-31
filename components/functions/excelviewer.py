from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QLineEdit,
    QTableWidget, QTableWidgetItem
)
from PyQt5.QtCore import pyqtSignal
from openpyxl import load_workbook
from components.functions import DropArea


class ExcelViewer(QWidget):
    columnsSelected = pyqtSignal(list)
    def __init__(self, drop_area):
        super().__init__()
        
        drop_area.fileDropped.connect(self.on_file)
        
        self.layout = QVBoxLayout()
        self.setLayout(self.layout)
        
        self.file_path_edit = QLineEdit()
        self.file_path_edit.setReadOnly(True)
        
        self.table_widget = QTableWidget()
        self.table_widget.setSelectionBehavior(QTableWidget.SelectColumns)
        self.table_widget.setSelectionMode(QTableWidget.MultiSelection)
        self.table_widget.itemSelectionChanged.connect(self.on_column_selection)
        
        # layout.add 
        self.layout.addWidget
        self.layout.addWidget(self.table_widget)
        
    def on_file(self, path):
        self.file_path = path
        self.file_path_edit.setText(path)
        self.load_excel_preview(path)
        
    def load_excel_preview(self, path):
        try:
            wb = load_workbook(path, read_only = False)
            sheet = wb.active
            self.table_widget.clear()
            self.table_widget.setRowCount(sheet.max_row)
            self.table_widget.setColumnCount(sheet.max_column)
            
            for r, row in enumerate(sheet.iter_rows(values_only=True)):
                for c, value in enumerate(row):
                    self.table_widget.setItem(
                        r, c, QTableWidgetItem(str(value)if value is not None else "")
                        )
                
            header_items = [sheet.cell(row=1, column=c+1).value for c in range(sheet.max_column)]
            self.table_widget.setHorizontalHeaderLabels([str(h) for h in header_items])
            
            wb.close()
            
        except Exception as e:
            print("Error loading Excel:", e)
    
    def open_excel_file(self):
        if not hasattr(self, 'file_path'):
            print("No file dropped yet.")
            return
        self.load_excel_preview(self.file_path)
        
    def on_column_selection(self):
        selected = {idx.column() for idx in self.table_widget.selectedIndexes()}
        selected = sorted(list(selected))
        
        if len(selected) == 2:
            self.columnsSelected.emit(selected)